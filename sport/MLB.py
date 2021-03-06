#!/usr/bin/env python
import Queue
import threading
import urllib, urllib2
import time, re
import optparse                          
from xlwt import *
import codecs
from BeautifulSoup import BeautifulSoup
from datetime import date
from punky import Browster
import os
from openpyxl import load_workbook
import datetime
from win32com.client import Dispatch

debugF = False

def weekDay():
    #print curDate.weekday()
    out = ''
    if curDate.weekday() == 0:
        out = 'M'
    elif curDate.weekday() == 1:
        out = 'TU'
    elif curDate.weekday() == 2:
        out = 'W'
    elif curDate.weekday() == 3:
        out = 'TH'
    elif curDate.weekday() == 4:
        out = 'F'
    elif curDate.weekday() == 5:
        out = 'SA'
    elif curDate.weekday() == 6:
        out = 'SU'
    return out
    #print weekDay
legues = {
"BAL":"AL",
"BOS":"AL",
"NYY":"AL",
"TB":"AL",
"TOR":"AL",
"CHW":"AL",
"CLE":"AL",
"DET":"AL",
"KC":"AL",
"MIN":"AL",
"LAA":"AL",
"OAK":"AL",
"SEA":"AL",
"TEX":"AL",
"ATL":"NL",
"MIA":"NL",
"NYM":"NL",
"PHI":"NL",
"WAS":"NL",
"CHC":"NL",
"CIN":"NL",
"HOU":"NL",
"MIL":"NL",
"PIT":"NL",
"ARI":"NL",
"COL":"NL",
"LA":"NL",
"SF":"NL",
"STL":"NL",
"SD":"NL"
}

mlbteams = {
'BAL':'2959',
'DET':'2978',
'TEX':'2977',
'TOR':'2984',
'CHW':'2974',
'KC' :'2965',
'TB' :'2960',
'LAA':'2979',
'CLE':'2980',
'OAK':'2969',
'MIN':'2983',
'SEA':'2973',
'BOS':'2966',
'NYY':'2970',
'CHC':'2982',
'CIN':'2961',
'LA' :'2967',
'ATL':'2957',
'NYM':'2964',
'WAS':'2972',
'ARI':'2968',
'HOU':'2981',
'PHI':'2958',
'MIL':'2976',
'PIT':'2971',
'STL':'2975',
'MIA':'2963',
'COL':'2956',
'SF' :'2962',
'SD' :'2955'
}

queue = Queue.Queue()

Dates = []
Opp = []
HV = []
WL = []
OU = []
class ThreadUrl(threading.Thread):
    """Threaded Url Grab"""
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue

    def run(self):
        while True:
            host = self.queue.get()
            #print host[host.rfind('/')+1:]
            #print 'MlbPastRes/' + host[host.find('pastresults/')+len('pastresults/'):host.rfind('/')] + '_' + host[host.rfind('/')+1:]
            if os.path.exists('MlbPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:]):
                vd = os.path.getmtime('MlbPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:])
                #print ' |vd :' + str(vd) + '| '
                vdd = datetime.date.fromtimestamp(vd)
                #print ' |vdd :' + str(vdd) + '| '
                #print vdd < curDate
                #print (vdd < curDate) or (not os.path.exists('pastres/' + host[host.rfind('/')+1:]))
                #print ' |curDate :' + str(curDate) + '| '
                #if not debugF:
                #str(curDate.year - years)
                if vdd < curDate:
                    urllib.urlretrieve(host, 'MlbPastRes/' + host[host.find('pastresults/')+len('pastresults/'):host.rfind('/')] + '_' + host[host.rfind('/')+1:])
            else:
                urllib.urlretrieve(host, 'MlbPastRes/' + host[host.find('pastresults/')+len('pastresults/'):host.rfind('/')] + '_' + host[host.rfind('/')+1:])
            self.queue.task_done()

start = time.time()
def main():

    for i in range(len(hosts)):
        t = ThreadUrl(queue)
        t.setDaemon(True)
        t.start()
    for host in hosts:
        queue.put(host)

    queue.join()

def getPastResults(teamName):
    
    #print 'Getting past results for ' + str(teamName)
    Dates = []
    Opp = []
    HV = []
    WL = []
    OU = []
    pastResLink = teamLinkHref[pastCount]
    #for years in range (0,2):
        
    #url = pastResLink[:pastResLink.rfind('/')] + '/pastresults/' + str(curDate.year - years) + pastResLink[pastResLink.rfind('/'):]
    #print nhlteams[teamName]
    for vYear in range (0,2):
        htmltext = open('MlbPastRes/' + str(curDate.year - vYear) + '_' + 'team' + mlbteams[teamName] + '.html','r')

        soup = BeautifulSoup(''.join(htmltext)) #parse html source
        
        #get past data
        tables = soup.findAll('table', attrs={'class':'data'}) # get required table
        for table in tables:
            if table.findAll('tr'):
                rows = table.findAll('tr') # get all rows in this table
                for tr in rows[1:]: 
                    cols = tr.findAll('td') # find all columns in table
                    
                    for td in cols[:1]: # Dates
                        DateTemp = map(int, str(td.find(text=True).strip()).split('/'))
                        vDate = date(int('20' + str(DateTemp[2]).zfill(2)), DateTemp[0], DateTemp[1])
                        Dates.append(vDate)
                        
                    for td in cols[1:2]: # HV
                        if td.find(text=True).replace('\n','').strip()[:1] == '@':
                            HV.append('V')
                        else:
                            HV.append('H')

                    for td in cols[1:2]: # Opp
                        Opp.append(td.find('a').find(text=True).strip())
                            
                    for td in cols[2:3]: # WL
                        WL.append(td.find(text=True).strip())
                        
                    for td in cols[6:7]: # OU
                        OU.append(td.find(text=True).replace('\n','').strip()[:1])
    #print [teamName, Dates, Opp, HV, WL, OU]
    return [teamName, Dates, Opp, HV, WL, OU]

def requestURL(vURL):
    #print 'Getting URL...'
    attempts = 0
    while attempts < 10:
        try:
            print vURL
            response = browser.load(vURL)
            attempts += 1
            break
        except:
            print 'Exception in address: ' + vURL
            print 'trying again after 4 sec...'
            response = ''
            time.sleep(4)
            attempts += 1
        
    #response
    return browser.html

#========================================================================

parser = optparse.OptionParser()
parser.add_option('-p', '--path', dest='path', help='path to csv') # date
parser.add_option('-d', '--date', dest='date', help='needed date') # date

options, args = parser.parse_args()

if not os.path.exists('MlbPastRes'):
    os.mkdir('MlbPastRes')

if not options.date: # date can be omited
    curDate = date.today()
else:
    
    DateTemp = map(int, str(options.date).split('/'))
    curDate = date(DateTemp[2], DateTemp[0], DateTemp[1])
    
months = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
myDate = date(2012,7,3)
#print (curDate - myDate).days
print curDate
#print curDate - datetime.timedelta(days=1)
if os.path.exists(options.path + 'res.csv'):
    os.remove(options.path + 'res.csv')

browser = Browster()
browser.create_webview(show=False)

if not debugF:
    storedhtml = requestURL('http://scores.covers.com/baseball-scores-matchups.aspx')
    soup = BeautifulSoup(''.join(storedhtml)) #parse html source

    # we will monitor this text. Once it's not present on the page - means ajax is done
    lookupPoint = soup.find('div', attrs={'class':'game-box'})['id']
    # lookupPoint
    #click calendar if current page is not for our day
    browser.runjs("javascript:ClientUpdateCalAndNav(5,'" + str(curDate.year) +"'," + str(curDate.year) +"," + str(curDate.month) +"," + str(curDate.day) +");")
    #<td onclick="javascript:ClientUpdateCalAndNav(5,'2012',2012,7,10);"><a href="javascript:void(0);" style="text-decoration:none;"><b>10</b></a></td>
    #print "javascript:ClientUpdateCalAndNav(5,'" + str(curDate.year) +"'," + str(curDate.year) +"," + str(curDate.month) +"," + str(curDate.day) +");"
        
    #while lookupPoint in browser.html:
        #print lookupPoint in browser.html
    #    browser.wait(0.2)
    #print lookupPoint in browser.html
    browser.wait(7)
# now we can get html source
if debugF:
    html = open(options.path + 'MLBScoreBoard.html','r')
else:
    html = browser.html
soup = BeautifulSoup(''.join(html)) #parse html source
#print soup

teamNamesArr = soup.findAll('td', attrs={'class':'datab'})
#print teamNamesArr
teamLinkHref= []
teamLinkName= []
SBhome = []
SBvis = []
j=0
for teamN in teamNamesArr:
    j+=1
    if j%2 == 0:
        if teamN.find('a', text=True).strip() not in SBhome:
            SBhome.append(teamN.find('a', text=True).strip())
    else:
        if teamN.find('a', text=True).strip() not in SBvis:
            SBvis.append(teamN.find('a', text=True).strip())
    if teamN.find('a', text=True).strip() not in teamLinkName:
        teamLinkHref.append(teamN.find('a')['href'])
        teamLinkName.append(teamN.find('a', text=True).strip())
#print teamLinkName


hosts = []
for tm in teamLinkName:
    #for years in range (0,2):
    #url = pastResLink[:pastResLink.rfind('/')] + '/pastresults/' + str(curDate.year - years) + pastResLink[pastResLink.rfind('/'):]
    hosts.append('http://www.covers.com/pageLoader/pageLoader.aspx?page=/data/mlb/teams/pastresults/' + str(curDate.year - 1) + '/team' + mlbteams[tm] + '.html') 
    hosts.append('http://www.covers.com/pageLoader/pageLoader.aspx?page=/data/mlb/teams/pastresults/' + str(curDate.year - 0) + '/team' + mlbteams[tm] + '.html') 
print hosts    
main()
#print "Elapsed Time: %s" % (time.time() - start)

'''
wb = load_workbook(filename = r'NHL.xlsx')
sheet_ranges = wb.get_sheet_by_name(name = 'Results')
crow = sheet_ranges.get_highest_row() + 2

sheet_ranges.cell('A'+str(crow)).value = str(options.date)
sheet_ranges.cell('A'+str(crow)).offset(2,0).value = weekDay()
sheet_ranges.cell('A'+str(crow)).offset(4,0).value = curDate.isocalendar()[1]
'''
vpath = str(options.path).replace('\\', '\\\\')

file_name = vpath + 'MLB.xlsx'
print file_name
excel = Dispatch('Excel.Application')
excel.Visible = False  #If we want to see it change, it's fun
workbook = excel.Workbooks.Open(file_name)
workBook = excel.ActiveWorkbook
sheets = workBook.Sheets
sheet = sheets('Results')
a_row = sheet.Range("A" + str(sheet.UsedRange.rows.Count)).End(-4162).Row + 2
c_row = sheet.Range("C" + str(sheet.UsedRange.rows.Count)).End(-4162).Row + 2
#print a_row, c_row
crow = max(a_row, c_row, 5)
#print crow
sheet.Activate()
#sheet_ranges.cell('A'+str(crow)).value = str(options.date)
#sheet_ranges.cell('A'+str(crow)).offset(2,0).value = weekDay()
#sheet_ranges.cell('A'+str(crow)).offset(4,0).value = curDate.isocalendar()[1]
sheet.Cells(crow,1).Value = str(curDate.month) + '/' + str(curDate.day) + '/' + str(curDate.year)
sheet.Cells(crow+2,1).Value = weekDay()
sheet.Cells(crow+4,1).Value = curDate.isocalendar()[1]

#print teamLinkHref
pastCount = 0
pastResults = []
for pastResName in teamLinkName:
    #return [teamName, Dates, Opp, HV, WL, OU]
    pastResults.append(getPastResults(pastResName))
    #pastResults.append([pastResName, Dates, Opp, HV, WL, OU])
    pastCount+=1
    #if pastCount ==2: break

#Q-T
pCount = 0
prevDate = curDate
for eachTeam in pastResults:
    vWc = 0
    tW = 0
    tL = 0
    i = 0
    vWLst = False
    for currWL in eachTeam[4]:
        if eachTeam[1][i] < curDate:
            #print currWL
            if vWc ==0:
                storedRes = currWL
            if storedRes == currWL:
                if currWL == 'W':
                    tW +=1
                elif currWL == 'L':
                    tL +=1
                storedRes = currWL
            else:
                break
            vWc+=1
        i+=1
    pCount +=1
    eachTeam.append(tW)
    eachTeam.append(tL)

    
    
    if pCount%2 == 0:
        #print pastResults[pCount-2][0]
        #print pastResults[pCount-1][0]
        
        #IJ
        if eachTeam[0] == pastResults[pCount-1][0]:
            IJO = 0
            IJU = 0
            IJst = False
            i = 0
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                #print res
                #print eachTeam[2][i] == pastResults[pCount-2][0]
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if not IJst:
                            storedres = res
                            IJst = True
                                
                        if storedres == res:
                            if str(res) == 'O':
                                IJO +=1
                            elif str(res) == 'U':
                                IJU +=1
                            storedres = res
                        else:
                            break
                        
                i+=1
            #print IJO, IJU 
        if IJO == 0: 
            IJO = ''
        if IJU == 0: 
            IJU = ''
            
        #KL
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            KLO = 0
            KLU = 0
            KLst = False
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if eachTeam[3][i] == 'H':
                            if not KLst:
                                storedres = res
                                KLst = True
                            if storedres == res:
                            
                                if str(res) == 'O':
                                    KLO +=1
                                elif str(res) == 'U':
                                    KLU +=1
                                storedres = res
                                
                            else:
                                break
                        
                i+=1
            #print KLO, KLU
        if KLO == 0: 
            KLO = ''
        if KLU == 0: 
            KLU = ''
            
        MNO = 0
        MNU = 0
        MNst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            MNCount = 0
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if str(res) == 'O':
                            MNO +=1
                        elif str(res) == 'U':
                            MNU +=1
                        MNCount+=1
                        if MNCount>=10:
                            break
                i+=1
            #print MNO, MNU    
        if MNO > MNU:
            MNU = 0
        elif MNO < MNU:
            MNO = 0
        if MNO == 0: 
            MNO = ''
        if MNU == 0: 
            MNU = ''
            
        vP = 0
        vPst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            
            i=0
            for res in eachTeam[2]:
                if eachTeam[1][i] <= curDate:
                    if not vPst:
                        prevres = pastResults[pCount-2][0]
                        vPst = True
                    if res == pastResults[pCount-2][0]:
                        if res == prevres:
                            vP+=1
                    
                    else:
                        break
                    prevres = res
                i+=1
#        print curDate
#        print weekDay()
        if vP == 0: 
            vP = ''

        #VW
        XYO = 0
        XYU = 0
        XYst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                #print i
                #print eachTeam[1][i]
                #print curDate
                if eachTeam[1][i] < curDate:
                    if not XYst:
                        storedres = res
                        XYst = True
                    #print storedres
                    #print res
                                    
                    if storedres == res:
                        if str(res) == 'O':
                            XYO +=1
                        elif str(res) == 'U':
                            XYU +=1
                        storedres = res
                        #print VWO, VWU
                    else:
                        break
                        
                i+=1
        #print XYO, XYU
        if XYO == 0: 
            XYO = ''
        if XYU == 0: 
            XYU = ''



        #AAAB
        ACADO = 0
        ACADU = 0
        ACADUst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            #print eachTeam[0]
            #print pastResults[pCount-1][0]
            i = 0
            for res in eachTeam[5]:
                #print res
                #print eachTeam[5]
                if eachTeam[1][i] < curDate:
                    #print eachTeam[1][i]
                    #print curDate
                    #print eachTeam[3][i]
                    if eachTeam[3][i] == 'H':
                        if not ACADUst:
                            storedres = res
                            ACADUst = True        
                        if storedres == res:
                            if str(res) == 'O':
                                ACADO +=1
                            elif str(res) == 'U':
                                ACADU +=1
                            storedres = res
                        else:
                            break
                        
                i+=1
        #print ACADO, ACADU
        if ACADO == 0: 
            ACADO = ''
        if ACADU == 0: 
            ACADU = ''


        if legues[str(pastResults[pCount-2][0])] == legues[str(pastResults[pCount-1][0])]:
            vG = legues[str(pastResults[pCount-1][0])][:1]
        else:
            vG = 'I'
        print vG
        
        output = [str(pastResults[pCount-2][0]), str(pastResults[pCount-1][0]), weekDay(), vG, 
                  str(IJO), str(IJU), str(KLO), str(KLU), str(MNO), str(MNU), str(vP),
                  str(0), str(pastResults[pCount-2][6]), str(pastResults[pCount-2][7]), 
                  str(pastResults[pCount-1][6]), str(pastResults[pCount-1][7]), 
                  str(0), str(VWO), str(VWU), str(XYO), str(XYU), str(0), str(AAABO), str(AAABU), str(ACADO), str(ACADU)]

        sheet.Range('C'+str(crow)).value = str(pastResults[pCount-2][0])
        sheet.Range('D'+str(crow)).value = str(pastResults[pCount-1][0])
        sheet.Range('E'+str(crow)).value = weekDay()
        sheet.Range('F'+str(crow)).value = vG
        #sheet.Range('G'+str(crow)).value = vG
        #sheet.Range('H'+str(crow)).value = vH
        
        sheet.Range('H'+str(crow)).value = str(IJO)
        sheet.Range('I'+str(crow)).value = str(IJU)
        sheet.Range('J'+str(crow)).value = str(KLO)
        sheet.Range('K'+str(crow)).value = str(KLU)
        sheet.Range('L'+str(crow)).value = str(MNO)
        sheet.Range('M'+str(crow)).value = str(MNU)
        
        sheet.Range('O'+str(crow)).value = vP
        
        if str(pastResults[pCount-2][6]) == str('0'):
            sheet.Range('P'+str(crow)).value = ''
        else:
            sheet.Range('P'+str(crow)).value = str(pastResults[pCount-2][6])
            
        if str(pastResults[pCount-2][7]) == str('0'):
            sheet.Range('Q'+str(crow)).value = ''
        else:
            sheet.Range('Q'+str(crow)).value = str(pastResults[pCount-2][7])
            
        if str(pastResults[pCount-1][6]) == str('0'):
            sheet.Range('R'+str(crow)).value = ''
        else:
            sheet.Range('R'+str(crow)).value = str(pastResults[pCount-1][6])
            
        if str(pastResults[pCount-1][7]) == str('0'):
            sheet.Range('S'+str(crow)).value = ''
        else:
            sheet.Range('S'+str(crow)).value = str(pastResults[pCount-1][7])
        
        sheet.Range('U'+str(crow)).value = str(VWO)
        sheet.Range('V'+str(crow)).value = str(VWU)
        sheet.Range('W'+str(crow)).value = str(XYO)
        sheet.Range('X'+str(crow)).value = str(XYU)
        
        sheet.Range('Z'+str(crow)).value = str(AAABO)
        sheet.Range('AA'+str(crow)).value = str(AAABU)
        sheet.Range('AB'+str(crow)).value = str(ACADO)
        sheet.Range('AC'+str(crow)).value = str(ACADU)
        
        crow +=1
        #wb.save(filename = r'NHL.xlsx')        
        print output
        #print str(pastResults[pCount-2][0]), str(pastResults[pCount-1][0]), str(IJO), str(IJU), str(KLO), str(KLU), str(MNO), str(MNU), str(vP), str(pastResults[pCount-2][6]), str(pastResults[pCount-2][7]), str(pastResults[pCount-1][6]), str(pastResults[pCount-1][7]), str(VWO), str(VWU), str(XYO), str(XYU), str(AAABO), str(AAABU), str(ACADO), str(ACADU)
        #print pastResults[pCount][0], pastResults[pCount+1][0], pastResults[pCount][6], pastResults[pCount][7], pastResults[pCount+1][6], pastResults[pCount+1][7]
    else:
    
                #VW
        #print eachTeam[0]
        #print pastResults[pCount-2][0]
        VWO = 0
        VWU = 0
        VWst = False
        #print eachTeam[0]
        #print pastResults[pCount-1][0]
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                #print res
                #print eachTeam[5]
                if eachTeam[1][i] < curDate:
                    if not VWst:
                        storedres = res
                        VWst = True        
                    if storedres == res:
                        if str(res) == 'O':
                            VWO +=1
                        elif str(res) == 'U':
                            VWU +=1
                        storedres = res
                    else:
                        break
                        
                i+=1
        #print VWO, VWU
        if VWO == 0: 
            VWO = ''
        if VWU == 0: 
            VWU = ''

        #ACAD
        #print eachTeam[0]
        #print pastResults[pCount-2][0]
        AAABO = 0
        AAABU = 0
        AAABUst = False
        #print eachTeam[0]
        #print pastResults[pCount-1][0]
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                    #print eachTeam[1][i]
                    #print eachTeam[0]
                    #print eachTeam[3][i]
                    if eachTeam[3][i] == 'V':
                        if not AAABUst:
                            storedres = res
                            AAABUst = True
                                        
                        if storedres == res:
                            if str(res) == 'O':
                                AAABO +=1
                            elif str(res) == 'U':
                                AAABU +=1
                            storedres = res
                        else:
                            break
                            
                i+=1
        #print AAABO, AAABU
        if AAABO == 0: 
            AAABO = ''
        if AAABU == 0: 
            AAABU = ''

workBook.Save
excel.Visible = True
#workBook.Close(SaveChanges=1)
#excel.Quit()
#excel.Visible = 1
#del excel

#!/usr/bin/env python
import Queue
import threading
import urllib, urllib2
import time, re
import optparse                          
from xlwt import *
import codecs
from BeautifulSoup import BeautifulSoup
from datetime import date
from punky import Browster
import os
from openpyxl import load_workbook
import datetime
from win32com.client import Dispatch
from decimal import Decimal

debugF = False

def weekDay():
    #print curDate.weekday()
    out = ''
    if curDate.weekday() == 0:
        out = 'M'
    elif curDate.weekday() == 1:
        out = 'TU'
    elif curDate.weekday() == 2:
        out = 'W'
    elif curDate.weekday() == 3:
        out = 'TH'
    elif curDate.weekday() == 4:
        out = 'F'
    elif curDate.weekday() == 5:
        out = 'SA'
    elif curDate.weekday() == 6:
        out = 'SU'
    return out
    #print weekDay
nhlteams = {
'COL':'1',
'DET':'2',
'BOS':'3',
'NYI':'4',
'CAR':'5',
'NYR':'6',
'PIT':'7',
'TOR':'8',
'OTT':'9',
'BUF':'10',
'MON':'11',
'PHI':'12',
'NJ':'13',
'FLA':'14',
'WAS':'15',
'STL':'16',
'TB':'17',
'CAL':'18',
'PHO':'19',
'DAL':'20',
'SJ':'21',
'CHI':'22',
'LA':'23',
'EDM':'24',
'ANA':'25',
'VAN':'26',
'NAS':'27',
'WIN':'28',
'MIN':'29',
'CLB':'30'
}

nhlteamsnames = {
'Colorado':        'COL',   
'Detroit':         'DET',   
'Boston':          'BOS',   
'NY Islanders':    'NYI',   
'Carolina':        'CAR',   
'NY Rangers':      'NYR',   
'Pittsburgh':      'PIT',   
'Toronto':         'TOR',   
'Ottawa':          'OTT',   
'Buffalo':         'BUF',   
'Montreal':        'MON',   
'Philadelphia':    'PHI',   
'New Jersey':      'NJ',    
'Florida':         'FLA',   
'Washington':      'WAS',   
'St. Louis':       'STL',   
'Tampa Bay':       'TB',    
'Calgary':         'CAL',   
'Phoenix':         'PHO',   
'Dallas':          'DAL',   
'San Jose':        'SJ',    
'Chicago':         'CHI',   
'Los Angeles':     'LA',    
'Edmonton':        'EDM',   
'Anaheim':         'ANA',   
'Vancouver':       'VAN',   
'Nashville':       'NAS',   
'Winnipeg':        'WIN',   
'Minnesota':       'MIN',   
'Columbus':        'CLB',
'Atlanta':        'WIN',
}   

legues = {
'COL':'W',
'DET':'W',
'BOS':'E',
'NYI':'E',
'CAR':'E',
'NYR':'E',
'PIT':'E',
'TOR':'E',
'OTT':'E',
'BUF':'E',
'MON':'E',
'PHI':'E',
'NJ':'E',
'FLA':'E',
'WAS':'E',
'STL':'W',
'TB':'E',
'CAL':'W',
'PHO':'W',
'DAL':'W',
'SJ':'W',
'CHI':'W',
'LA':'W',
'EDM':'W',
'ANA':'W',
'VAN':'W',
'NAS':'W',
'WIN':'E',
'MIN':'W',
'CLB':'W'
}
queue = Queue.Queue()

Dates = []
Opp = []
HV = []
WL = []
OU = []
class ThreadUrl(threading.Thread):
    """Threaded Url Grab"""
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue

    def run(self):
        while True:
            host = self.queue.get()
            #print host[host.rfind('/')+1:]
            if os.path.exists('NhlPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:]):
                vd = os.path.getmtime('NhlPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:])
                #print ' |vd :' + str(vd) + '| '
                vdd = datetime.date.fromtimestamp(vd)
                #print ' |vdd :' + str(vdd) + '| '
                #print vdd < curDate
                #print (vdd < curDate) or (not os.path.exists('pastres/' + host[host.rfind('/')+1:]))
                #print ' |curDate :' + str(curDate) + '| '
                #if not debugF:
                if vdd < curDate:
                    urllib.urlretrieve(host, 'NhlPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:])
            else:
                urllib.urlretrieve(host, 'NhlPastRes/' + host[host.rfind('-')+1:host.rfind('-')+5] + '_' + host[host.rfind('/')+1:])
            self.queue.task_done()

start = time.time()
def main():

    for i in range(len(hosts)):
        t = ThreadUrl(queue)
        t.setDaemon(True)
        t.start()
    for host in hosts:
        queue.put(host)

    queue.join()

def getPastResults(teamName):
    
    #print 'Getting past results for ' + str(teamName)
    Dates = []
    Opp = []
    HV = []
    WL = []
    OU = []
    pastResLink = teamLinkHref[pastCount]
    #for years in range (0,2):
        
    #url = pastResLink[:pastResLink.rfind('/')] + '/pastresults/' + str(curDate.year - years) + pastResLink[pastResLink.rfind('/'):]
    #print nhlteams[teamName]
    for vYear in range (0,2):
        htmltext = open('NhlPastRes/' + str(curDate.year - vYear) + '_' + 'team' + nhlteams[teamName] + '.html','r')

        soup = BeautifulSoup(''.join(htmltext)) #parse html source
        
        #get past data
        tables = soup.findAll('table', attrs={'class':'data'}) # get required table
        for table in tables:
            if table.findAll('tr'):
                rows = table.findAll('tr') # get all rows in this table
                for tr in rows[1:]: 
                    cols = tr.findAll('td') # find all columns in table
                    
                    for td in cols[:1]: # Dates
                        DateTemp = map(int, str(td.find(text=True).strip()).split('/'))
                        vDate = date(int('20' + str(DateTemp[2]).zfill(2)), DateTemp[0], DateTemp[1])
                        Dates.append(vDate)
                        
                    for td in cols[1:2]: # HV
                        if td.find(text=True).replace('\n','').strip()[:1] == '@':
                            HV.append('V')
                        else:
                            HV.append('H')

                    for td in cols[1:2]: # Opp
                        Opp.append(nhlteamsnames[td.find('a').find(text=True).strip()])
                            
                    for td in cols[2:3]: # WL
                        WL.append(td.find(text=True).strip())
                        
                    for td in cols[6:7]: # OU
                        OU.append(td.find(text=True).replace('\n','').strip()[:1])
    #print [teamName, Dates, Opp, HV, WL, OU]
    return [teamName, Dates, Opp, HV, WL, OU]

def requestURL(vURL):
	#print 'Getting URL...'
	attempts = 0
	while attempts < 10:
		try:
			print vURL
			response = browser.load(vURL)
			attempts += 1
			break
		except:
			print 'Exception in address: ' + vURL
			print 'trying again after 4 sec...'
			response = ''
			time.sleep(4)
			attempts += 1
		
	#response
	return browser.html

#========================================================================

parser = optparse.OptionParser()
parser.add_option('-p', '--path', dest='path', help='path to csv') # date
parser.add_option('-d', '--date', dest='date', help='needed date') # date

options, args = parser.parse_args()

if not os.path.exists('NhlPastRes'):
    os.mkdir('NhlPastRes')

if not options.date: # date can be omited
    curDate = date.today()
else:
    
    DateTemp = map(int, str(options.date).split('/'))
    curDate = date(DateTemp[2], DateTemp[0], DateTemp[1])
    
months = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
myDate = date(2012,7,3)
#print (curDate - myDate).days
print curDate
#print curDate - datetime.timedelta(days=1)
if os.path.exists(options.path + 'res.csv'):
    os.remove(options.path + 'res.csv')

browser = Browster()
browser.create_webview(show=False)

if not debugF:
    storedhtml = requestURL('http://scores.covers.com/hockey-scores-matchups.aspx')
    soup = BeautifulSoup(''.join(storedhtml)) #parse html source

    # we will monitor this text. Once it's not present on the page - means ajax is done
    lookupPoint = soup.find('div', attrs={'class':'game-box'})['id']
    # lookupPoint
    #click calendar if current page is not for our day
    browser.runjs("javascript:ClientUpdateCalAndNav(6,'" + str(curDate.year) + '-' + str(curDate.year + 1) + "'," + str(curDate.year) +"," + str(curDate.month) +"," + str(curDate.day) +");")
    #<td onclick="javascript:ClientUpdateCalAndNav(5,'2012',2012,7,10);"><a href="javascript:void(0);" style="text-decoration:none;"><b>10</b></a></td>
    #print "javascript:ClientUpdateCalAndNav(5,'" + str(curDate.year) +"'," + str(curDate.year) +"," + str(curDate.month) +"," + str(curDate.day) +");"
        
    #while lookupPoint in browser.html:
        #print lookupPoint in browser.html
    #    browser.wait(0.2)
    #print lookupPoint in browser.html
    browser.wait(7)
# now we can get html source
if debugF:
    html = open(options.path + 'NHLScoreBoard.html','r')
else:
    html = browser.html
soup = BeautifulSoup(''.join(html)) #parse html source
#print soup

teamNamesArr = soup.findAll('td', attrs={'class':'datab'})

teamLinkHref= []
teamLinkName= []
SBhome = []
SBvis = []
vfav = {}
j=0
for teamN in teamNamesArr:
    j+=1
    if j%2 == 0:
        if teamN.find('a', text=True).strip() not in SBhome:
            SBhome.append(teamN.find('a', text=True).strip())
            vfav.update({teamN.find('a', text=True).strip():teamN.previous.previous.findAll('td')[-1].text.strip()})
    else:
        if teamN.find('a', text=True).strip() not in SBvis:
            SBvis.append(teamN.find('a', text=True).strip())
    if teamN.find('a', text=True).strip() not in teamLinkName:
        teamLinkHref.append(teamN.find('a')['href'])
        teamLinkName.append(teamN.find('a', text=True).strip())

hosts = []
for tm in teamLinkName:
    #for years in range (0,2):
    hosts.append('http://www.covers.com/pageLoader/pageLoader.aspx?page=/data/nhl/teams/pastresults/' + str(curDate.year - 2) + '-' + str(curDate.year - 1)  + '/team' + nhlteams[tm] + '.html') 
    hosts.append('http://www.covers.com/pageLoader/pageLoader.aspx?page=/data/nhl/teams/pastresults/' + str(curDate.year - 1) + '-' + str(curDate.year - 0)  + '/team' + nhlteams[tm] + '.html') 
print hosts    
main()
#print "Elapsed Time: %s" % (time.time() - start)

'''
wb = load_workbook(filename = r'NHL.xlsx')
sheet_ranges = wb.get_sheet_by_name(name = 'Results')
crow = sheet_ranges.get_highest_row() + 2

sheet_ranges.cell('A'+str(crow)).value = str(options.date)
sheet_ranges.cell('A'+str(crow)).offset(2,0).value = weekDay()
sheet_ranges.cell('A'+str(crow)).offset(4,0).value = curDate.isocalendar()[1]
'''
vpath = str(options.path).replace('\\', '\\\\')

file_name = vpath + 'NHL.xlsx'
print file_name
excel = Dispatch('Excel.Application')
excel.Visible = False  #If we want to see it change, it's fun
workbook = excel.Workbooks.Open(file_name)
workBook = excel.ActiveWorkbook
sheets = workBook.Sheets
sheet = sheets('Results')
a_row = sheet.Range("A" + str(sheet.UsedRange.rows.Count)).End(-4162).Row + 2
c_row = sheet.Range("C" + str(sheet.UsedRange.rows.Count)).End(-4162).Row + 2
#print a_row, c_row
crow = max(a_row, c_row, 5)
#print crow
sheet.Activate()
#sheet_ranges.cell('A'+str(crow)).value = str(options.date)
#sheet_ranges.cell('A'+str(crow)).offset(2,0).value = weekDay()
#sheet_ranges.cell('A'+str(crow)).offset(4,0).value = curDate.isocalendar()[1]
sheet.Cells(crow,1).Value = str(curDate.month) + '/' + str(curDate.day) + '/' + str(curDate.year)
sheet.Cells(crow+2,1).Value = weekDay()
sheet.Cells(crow+4,1).Value = curDate.isocalendar()[1]

#print teamLinkHref
pastCount = 0
pastResults = []
for pastResName in teamLinkName:
    #return [teamName, Dates, Opp, HV, WL, OU]
    pastResults.append(getPastResults(pastResName))
    #pastResults.append([pastResName, Dates, Opp, HV, WL, OU])
    pastCount+=1
    #if pastCount ==2: break

#Q-T
pCount = 0
prevDate = curDate
for eachTeam in pastResults:
    #print eachTeam
    vWc = 0
    tW = 0
    tL = 0
    i = 0
    vWLst = False
    for currWL in eachTeam[4]:
        if eachTeam[1][i] < curDate:
            #print currWL
            if vWc ==0:
                storedRes = currWL
            if storedRes == currWL:
                if currWL == 'W':
                    tW +=1
                elif currWL == 'L':
                    tL +=1
                storedRes = currWL
            else:
                break
            vWc+=1
        i+=1

    pCount +=1
    eachTeam.append(tW)
    eachTeam.append(tL)

    
    
    if pCount%2 == 0:
        #print pastResults[pCount-2][0]
        #print pastResults[pCount-1][0]
        
        #IJ
        if eachTeam[0] == pastResults[pCount-1][0]:
            #print 'here'
            IJO = 0
            IJU = 0
            IJst = False
            i = 0
            for res in eachTeam[5]:
                #print res
                #print eachTeam[1][i], curDate
                #print eachTeam[1][i] < curDate
                
                if eachTeam[1][i] < curDate:
                    #print res
                    #print eachTeam[2][i] 
                    #print pastResults[pCount-2][0]
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if not IJst:
                            storedres = res
                            IJst = True
                            #print 'here1'
                                
                        if storedres == res:
                            if str(res) == 'O':
                                IJO +=1
                            elif str(res) == 'U':
                                IJU +=1
                            storedres = res
                        else:
                            break
                        
                i+=1
            #print IJO, IJU 
        if IJO == 0: 
            IJO = ''
        if IJU == 0: 
            IJU = ''
            
        #KL
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            KLO = 0
            KLU = 0
            KLst = False
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if eachTeam[3][i] == 'H':
                            if not KLst:
                                storedres = res
                                KLst = True
                            if storedres == res:
                            
                                if str(res) == 'O':
                                    KLO +=1
                                elif str(res) == 'U':
                                    KLU +=1
                                storedres = res
                                
                            else:
                                break
                        
                i+=1
            #print KLO, KLU
        if KLO == 0: 
            KLO = ''
        if KLU == 0: 
            KLU = ''
            
        MNO = 0
        MNU = 0
        MNst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            MNCount = 0
            for res in eachTeam[5]:
                #print eachTeam[1][i], eachTeam[2][i], eachTeam[3][i], eachTeam[4][i], eachTeam[5][i]
                #print curDate
                if eachTeam[1][i] < curDate:
                    if eachTeam[2][i] == pastResults[pCount-2][0]:
                        if str(res) == 'O':
                            MNO +=1
                        elif str(res) == 'U':
                            MNU +=1
                        MNCount+=1
                        if MNCount>=5:
                            break
                i+=1
            #print MNO, MNU    
        if MNO > MNU:
            MNU = 0
        elif MNO < MNU:
            MNO = 0
        if MNO == 0: 
            MNO = ''
        if MNU == 0: 
            MNU = ''

        vGV = 0
        if eachTeam[0] == pastResults[pCount-1][0]:
            i=0
            for res in eachTeam[2]:
                #print pastResults[pCount-1][1]
                #print date(curDate.year, curDate.month, curDate.day - 1)
                #print len(pastResults[pCount-1][1])
                if curDate - datetime.timedelta(days=1) in pastResults[pCount-1][1]:
                    #print pastResults[pCount-1][1].index(date(curDate.year, curDate.month, curDate.day - 1))
                    vGV = 1
                i+=1
        if vGV == 0: 
            vGV = ''

                
        vP = 0
        vPst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            
            i=0
            for res in eachTeam[2]:
                if eachTeam[1][i] <= curDate:
                    if not vPst:
                        prevres = pastResults[pCount-2][0]
                        vPst = True
                    #print pastResults[pCount-2][0]
                    if res == pastResults[pCount-2][0]:
                        if res == prevres:
                            vP+=1
                    
                    else:
                        break
                    prevres = res
                i+=1
#        print curDate
#        print weekDay()
        if vP == 0: 
            vP = ''

        #VW
        XYO = 0
        XYU = 0
        XYst = False
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                #print i
                #print eachTeam[1][i]
                #print curDate
                if eachTeam[1][i] < curDate:
                    if not XYst:
                        storedres = res
                        XYst = True
                    #print storedres
                    #print res
                                    
                    if storedres == res:
                        if str(res) == 'O':
                            XYO +=1
                        elif str(res) == 'U':
                            XYU +=1
                        storedres = res
                        #print VWO, VWU
                    else:
                        break
                        
                i+=1
        #print XYO, XYU
        if XYO == 0: 
            XYO = ''
        if XYU == 0: 
            XYU = ''



        #AAAB
        ACADO = 0
        ACADU = 0
        ACADUst = False
        vHH = 0
        vHHst = True
        if eachTeam[0] == pastResults[pCount-1][0]:
            #print eachTeam[0]
            #print pastResults[pCount-1][0]
            i = 0
            for res in eachTeam[5]:
                #print res
                #print eachTeam[5]
                if eachTeam[1][i] < curDate:
                    #print eachTeam[1][i]
                    #print curDate
                    #print eachTeam[3][i]
                    if eachTeam[3][i] == 'H':
                        if vHHst:
                            vHH = 1
                            vHHst = False
                        if not ACADUst:
                            storedres = res
                            ACADUst = True        
                        if storedres == res:
                            if str(res) == 'O':
                                ACADO +=1
                            elif str(res) == 'U':
                                ACADU +=1
                            storedres = res
                        else:
                            break
                    else:
                        if vHHst:
                            vHH = 0
                            vHHst = False

                i+=1
        if ACADO == 0: 
            ACADO = ''
        if ACADU == 0: 
            ACADU = ''

        #print ACADO, ACADU
        if legues[str(pastResults[pCount-2][0])] == legues[str(pastResults[pCount-1][0])]:
            vF = legues[str(pastResults[pCount-1][0])]
        else:
            vF = 'I'
        #print vF
        
        if vGV == 1 and vGH == 1:
            vG = 'VH'
        elif vGV == 1:
            vG = 'V'
        elif vGH == 1:
            vG = 'H'
        else:
            vG = ''
        #print vG
        
        if vVV == 1 and vHH == 1:
            vH = 'HV'
        elif vVV == 1 and vHH == 0:
            vH = 'HH'
        elif vVV == 0 and vHH == 0:
            vH = 'VH'
        elif vVV == 0 and vHH == 1:
            vH = 'VV'

        output = [str(pastResults[pCount-2][0]), str(pastResults[pCount-1][0]), weekDay(), vF, vG, vH, 
                  str(IJO), str(IJU), str(KLO), str(KLU), str(MNO), str(MNU), 
                  str(0), str(pastResults[pCount-2][6]), str(pastResults[pCount-2][7]), 
                  str(pastResults[pCount-1][6]), str(pastResults[pCount-1][7]), 
                  str(0), str(VWO), str(VWU), str(XYO), str(XYU), str(0), str(AAABO), str(AAABU), str(ACADO), str(ACADU)]

        sheet.Range('C'+str(crow)).value = str(pastResults[pCount-2][0])
        sheet.Range('D'+str(crow)).value = str(pastResults[pCount-1][0])
        sheet.Range('E'+str(crow)).value = weekDay()
        try:
            if Decimal(vfav[str(pastResults[pCount-1][0])]) >= 0:
                sheet.Range('F'+str(crow)).value = 'V'
            elif Decimal(vfav[str(pastResults[pCount-1][0])]) < 0:
                sheet.Range('F'+str(crow)).value = 'H'
        except:
            sheet.Range('F'+str(crow)).value = ''
        sheet.Range('G'+str(crow)).value = vF
        sheet.Range('H'+str(crow)).value = vG
        sheet.Range('I'+str(crow)).value = vH
        
        sheet.Range('K'+str(crow)).value = str(IJO)
        sheet.Range('L'+str(crow)).value = str(IJU)
        sheet.Range('M'+str(crow)).value = str(KLO)
        sheet.Range('N'+str(crow)).value = str(KLU)
        sheet.Range('O'+str(crow)).value = str(MNO)
        sheet.Range('P'+str(crow)).value = str(MNU)

        
        if str(pastResults[pCount-2][6]) == str('0'):
            sheet.Range('R'+str(crow)).value = ''
        else:
            sheet.Range('R'+str(crow)).value = str(pastResults[pCount-2][6])
            
        if str(pastResults[pCount-2][7]) == str('0'):
            sheet.Range('S'+str(crow)).value = ''
        else:
            sheet.Range('S'+str(crow)).value = str(pastResults[pCount-2][7])
            
        if str(pastResults[pCount-1][6]) == str('0'):
            sheet.Range('T'+str(crow)).value = ''
        else:
            sheet.Range('T'+str(crow)).value = str(pastResults[pCount-1][6])
            
        if str(pastResults[pCount-1][7]) == str('0'):
            sheet.Range('U'+str(crow)).value = ''
        else:
            sheet.Range('U'+str(crow)).value = str(pastResults[pCount-1][7])
        
        
        #sheet.Range('Q'+str(crow)).value = str(pastResults[pCount-2][6])
        #sheet.Range('R'+str(crow)).value = str(pastResults[pCount-2][7])
        #sheet.Range('S'+str(crow)).value = str(pastResults[pCount-1][6])
        #sheet.Range('T'+str(crow)).value = str(pastResults[pCount-1][7])
        
        sheet.Range('W'+str(crow)).value = str(VWO)
        sheet.Range('X'+str(crow)).value = str(VWU)
        sheet.Range('Y'+str(crow)).value = str(XYO)
        sheet.Range('Z'+str(crow)).value = str(XYU)
        
        sheet.Range('AA'+str(crow)).value = str(AAABO)
        sheet.Range('AB'+str(crow)).value = str(AAABU)
        sheet.Range('AC'+str(crow)).value = str(ACADO)
        sheet.Range('AD'+str(crow)).value = str(ACADU)
        
        crow +=1
        #wb.save(filename = r'NHL.xlsx')        
        print output
        #print str(pastResults[pCount-2][0]), str(pastResults[pCount-1][0]), str(IJO), str(IJU), str(KLO), str(KLU), str(MNO), str(MNU), str(vP), str(pastResults[pCount-2][6]), str(pastResults[pCount-2][7]), str(pastResults[pCount-1][6]), str(pastResults[pCount-1][7]), str(VWO), str(VWU), str(XYO), str(XYU), str(AAABO), str(AAABU), str(ACADO), str(ACADU)
        #print pastResults[pCount][0], pastResults[pCount+1][0], pastResults[pCount][6], pastResults[pCount][7], pastResults[pCount+1][6], pastResults[pCount+1][7]
    else:
    
        vGH = 0
        if eachTeam[0] == pastResults[pCount-1][0]:
            i=0
            for res in eachTeam[2]:
                
                if curDate - datetime.timedelta(days=1) in pastResults[pCount-1][1]:
                    vGH = 1
                i+=1
        if vGH == 0: 
            vGH = ''

                #VW
        #print eachTeam[0]
        #print pastResults[pCount-2][0]
        VWO = 0
        VWU = 0
        VWst = False
        #print eachTeam[0]
        #print pastResults[pCount-1][0]
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                #print res
                #print eachTeam[5]
                if eachTeam[1][i] < curDate:
                    if not VWst:
                        storedres = res
                        VWst = True        
                    if storedres == res:
                        if str(res) == 'O':
                            VWO +=1
                        elif str(res) == 'U':
                            VWU +=1
                        storedres = res
                    else:
                        break
                        
                i+=1
        #print VWO, VWU
        if VWO == 0: 
            VWO = ''
        if VWU == 0: 
            VWU = ''

        #ACAD
        #print eachTeam[0]
        #print pastResults[pCount-2][0]
        AAABO = 0
        AAABU = 0
        AAABUst = False
        vVV = 0
        vVVst = True
        #print eachTeam[0]
        #print pastResults[pCount-1][0]
        if eachTeam[0] == pastResults[pCount-1][0]:
            i = 0
            for res in eachTeam[5]:
                if eachTeam[1][i] < curDate:
                    #print eachTeam[1][i]
                    #print eachTeam[0]
                    #print eachTeam[3][i]
                    if eachTeam[3][i] == 'V':
                        if vVVst:
                            vVV = 1
                            vVVst = False
                        if not AAABUst:
                            storedres = res
                            AAABUst = True
                                        
                        if storedres == res:
                            if str(res) == 'O':
                                AAABO +=1
                            elif str(res) == 'U':
                                AAABU +=1
                            storedres = res
                        else:
                            break
                    else:
                        if vVVst:
                            vVV = 0
                            vVVst = False
                            
                i+=1
        #print AAABO, AAABU
        if AAABO == 0: 
            AAABO = ''
        if AAABU == 0: 
            AAABU = ''

workBook.Save
excel.Visible = True
#workBook.Close(SaveChanges=1)
#excel.Quit()
#excel.Visible = 0 
#del excel

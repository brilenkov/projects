from openpyxl import load_workbook
import Queue
import threading
import urllib, urllib2
import os
import csv
from pythonutils import OrderedDict
import datetime
from datetime import date
import time
from BeautifulSoup import BeautifulSoup

def nextWorkDay(cd, incr):
    if (cd + datetime.timedelta(days=incr)).weekday() == 5:
        return cd + datetime.timedelta(days=incr+2)
    elif (cd + datetime.timedelta(days=incr)).weekday() == 6:
        return cd + datetime.timedelta(days=incr+1)
    else:
        return cd + datetime.timedelta(days=incr)

debugF = False
#debugF = True
logfile = open('%s_%s.txt' % ('StocksLog', int(time.time())), 'a')    
def printAndLog(line):
    try:
        logfile.write(str(line))
    except:
        logfile.write(str(''.join(line)))
        logfile.write('\n')
    print line
    return 

queue = Queue.Queue()
queue1 = Queue.Queue()

months = {
'Jan':1,
'Feb':2,
'Mar':3,
'Apr':4,
'May':5,
'Jun':6,
'Jul':7,
'Aug':8,
'Sep':9,
'Oct':10,
'Nov':11,
'Dec':12
}

if not os.path.exists('Stocks'):
    os.mkdir('Stocks')
for f in os.listdir('./Stocks'):
    os.remove('./Stocks/' + f)

class ThreadUrl(threading.Thread):
    """Threaded Url Grab"""
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue
    def run(self):
        while True:
            sn = self.queue.get()
            #http://finance.yahoo.com/q/hp?s=CGR
            urllib.urlretrieve('http://ichart.finance.yahoo.com/table.csv?s=' + sn, 'Stocks/' + sn + '.csv')
            self.queue.task_done()
            
class ThreadUrlStocks(threading.Thread):
    """Threaded Url Grab"""
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue

    def run(self):
        while True:
            vname = self.queue.get()
            attempts = 0
            while attempts < 10: 
                try:
                    #printAndLog('http://finance.yahoo.com/q/hp?s=' + str(vname[:-4]).strip() + '+Historical+Prices')
                    html = urllib2.urlopen('http://finance.yahoo.com/q/hp?s=' + str(vname[:-4]).strip() + '+Historical+Prices', None)
                    attempts += 1
                    break
                except urllib2.HTTPError, error: 
                    printAndLog("Connection ERROR. Reconnecting after 4 sec...")
                    html = ''
                    time.sleep(4)
                    attempts += 1
                    
                printAndLog('*** Unable to connect after ' + str(attempts) + ' tries!')
            
            os.remove('./Stocks/' + str(vname))
            newcsv = open('./Stocks/' + vname, 'wb')
            w = csv.writer(newcsv, dialect='excel')    
            w.writerow(['Date','Open','High','Low','Close','Volume','Adj Close'])
            
            soup = BeautifulSoup(''.join(html))
            try:
                StocksTable = soup.find('table', attrs={'class':'yfnc_datamodoutline1'}).find('table')
            
                rows = StocksTable.findAll('tr') 
                for tr in rows: 
                    stockData = []
                    cols = tr.findAll('td') 
                    #'Date','Open','High','Low','Close','Volume','Adj Close'
                    vpass = False
                    for td in cols[:1]:
                        stockText = td.find(text=True)
                        stockArr = stockText.replace(',','').split(' ')
                        #printAndLog(len(stockArr)
                        if len(stockArr) == 3:
                            vpass = True
                            stockData.append(str(stockArr[2]) + '-' + str(months[stockArr[0]]).zfill(2) + '-' +  str(stockArr[1]).zfill(2))
                    if vpass:
                        for td in cols[1:]:
                            stockData.append(td.find(text=True).replace(',',''))
                        #printAndLog(stockData       
                        w.writerow(stockData)
                #self.queue.task_done()
            except:
                printAndLog('*** ERROR: web page does not contain required table for "' + str(vname[:-4]) + '" stock!')
                newcsv.close()
                os.remove('./Stocks/' + str(vname))
                #self.queue.task_done()
                #break
                pass
            newcsv.close()
            self.queue.task_done()
wb = load_workbook(filename = r'Stocks.xlsx')            
if not debugF:
    printAndLog('DEBUG: parsing Stocks.xlsx to get all stock names')
    #wb = load_workbook(filename = r'Stocks.xlsx', optimized_write = True)
    
    #wb = Workbook(optimized_write = True)
    stocknames = []
    count = 0
    for sheet in wb.get_sheet_names():
        
        count+=1
        ws = wb.get_sheet_by_name(sheet)
        
        #printAndLog(sheet)
        for row in range(1,ws.get_highest_row()+1):
            if ws.cell('C' + str(row)).value is not None and 'Ticker' not in ws.cell('C' + str(row)).value and ws.cell('R' + str(row)).value is None and ws.cell('S' + str(row)).value is None and ws.cell('T' + str(row)).value is None and ws.cell('U' + str(row)).value is None and ws.cell('V' + str(row)).value is None and ws.cell('W' + str(row)).value is None:
                stocknames.append(ws.cell('C' + str(row)).value)
                
    stocknamesset = set(stocknames)
    printAndLog('DEBUG: found ' + str(len(stocknamesset)) + ' unique stock names')
    printAndLog('DEBUG: downloading csv files...')
    #for i in range(len(stocknamesset)):
    for i in range(0,50):
        t = ThreadUrl(queue)
        t.setDaemon(True)
        t.start()
    for vset in stocknamesset:
        #printAndLog(vset)
        queue.put(vset)
    queue.join()
    
    
printAndLog('DEBUG: opening csv files and get data')
stocksdict = {}
badfiles = []

for f in os.listdir('./Stocks'):
    #printAndLog(f)
    r = csv.reader(open('./Stocks/' + f, 'rb'), dialect='excel')
    datadict = {} #OrderedDict()
    rc = 0
    namegood = True
    for row in r:
        rc+=1
        if rc > 1:
            try:
                DateTemp = map(int, str(row[0]).split('-'))
                mDate = datetime.datetime(DateTemp[0], DateTemp[1], DateTemp[2], 0, 0)
                datadict[mDate] = [row[1], row[6]]
                
            except:
                printAndLog('*** WARNING: Current file "' + str(f) + '" seems to be wrong... ')
                badfiles.append(f)
                namegood = False
                #stocksdict.pop(f[:-4])
                break
    if namegood:
        stocksdict.update({f[:-4]:datadict})

skipname = []    
if badfiles:
    for i in range(len(badfiles)):
        t = ThreadUrlStocks(queue1)
        t.setDaemon(True)
        t.start()
    for badfile in badfiles:
        printAndLog('DEBUG: trying to get stocks for "' + str(badfile[:-4]) + '" from site again')
        queue1.put(badfile)
    queue1.join()
    
    for badfile in badfiles:
        vbreak = False
        try:
            r = csv.reader(open('./Stocks/' + badfile, 'rb'), dialect='excel')
            datadict = {} #OrderedDict()
            rc = 0
            for row in r:
                rc+=1
                if rc > 1:
                    try:
                        DateTemp = map(int, str(row[0]).split('-'))
                        mDate = datetime.datetime(DateTemp[0], DateTemp[1], DateTemp[2], 0, 0)
                        datadict[mDate] = [row[1], row[6]]
                        #badfiles.pop(f)
                    except:
                        printAndLog('*** ERROR: Unable to get data for  "' + str(badfile[:-4]) + '" stock!!!')
                        vbreak = True
                        break
            if not vbreak:
                stocksdict.update({badfile[:-4]:datadict})
        except:
            skipname.append(badfile[:-4])
            pass
printAndLog('*** WARNING: following stock names will be skipped:')
printAndLog(skipname)

#wb = load_workbook(filename = r'Stocks.xlsx')
printAndLog('DEBUG: fill Stocks.xlsx')

for sheet in wb.get_sheet_names():
    ws = wb.get_sheet_by_name(sheet)
    vpass = True
    printAndLog('DEBUG: current sheet: "' + str(sheet) + '"')
    for row in range(1,ws.get_highest_row()+1):
        c = ws.cell('C' + str(row)).value
        if str(c) not in skipname:
            if c is not None and 'Ticker' in str(c):
                if ws.cell('A' + str(row)).value is not None:
                    curdate = ws.cell('A' + str(row)).value
                    vpass = True
                elif ws.cell('B' + str(row)).value is not None:
                    curdate = ws.cell('B' + str(row)).value
                    vpass = True
                else:
                    vpass = False
                    
                    
            if vpass:    
                #print c
                #print curdate
                #print curdate + datetime.timedelta(days=1)
                #try:
                #    print stocksdict[c]
                #except:
                #    pass
                if c is not None and 'Ticker' not in str(c):
                    if ws.cell('R' + str(row)).value is None or ws.cell('S' + str(row)).value is None or ws.cell('T' + str(row)).value is None or ws.cell('U' + str(row)).value is None:
                        #try:
                            #print stocksdict[c]
                            #print stocksdict[c][curdate + datetime.timedelta(days=1)]
                            #print stocksdict[c][curdate + datetime.timedelta(days=1)].values()
                            #print stocksdict[c][curdate + datetime.timedelta(days=1)].values()[0][0]
                            try:
                                #print stocksdict[c][curdate + datetime.timedelta(days=1)].values()[0][0]
                                #print stocksdict[c][stocksdict[c].index(curdate)-1:stocksdict[c].index(curdate)].values()[0][1]
#                                ws.cell('R' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-1:stocksdict[c].index(curdate)].values()[0][0]
#                                ws.cell('S' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-1:stocksdict[c].index(curdate)].values()[0][1]
                                ws.cell('R' + str(row)).value = stocksdict[c][nextWorkDay(curdate,1)][0]
                                ws.cell('S' + str(row)).value = stocksdict[c][nextWorkDay(curdate,1)][1]
                                try:
#                                    ws.cell('T' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-2:stocksdict[c].index(curdate)-1].values()[0][0]
#                                    ws.cell('U' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-2:stocksdict[c].index(curdate)-1].values()[0][1]
                                    ws.cell('T' + str(row)).value = stocksdict[c][nextWorkDay(curdate,2)][0]
                                    ws.cell('U' + str(row)).value = stocksdict[c][nextWorkDay(curdate,2)][1]
                                    try:
#                                        ws.cell('V' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-5:stocksdict[c].index(curdate)-4].values()[0][0]
#                                        ws.cell('W' + str(row)).value = stocksdict[c][stocksdict[c].index(curdate)-5:stocksdict[c].index(curdate)-4].values()[0][1]
                                        ws.cell('V' + str(row)).value = stocksdict[c][nextWorkDay(curdate,7)][0]
                                        ws.cell('W' + str(row)).value = stocksdict[c][nextWorkDay(curdate,7)][1]
                                    except:
                                        printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the 7-th day after. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                        pass
                                except:
                                    printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the day after tomorrow. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                    printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the 7-th day after. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                    pass
                            except:
                                printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the next day. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the day after tomorrow. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                printAndLog('*** ERROR: Line ' + str(row) + '. No data to fill for the 7-th day after. See sheet "' + str(sheet) + '" and file "' + str(c) + '.csv" ')
                                pass
                                
                        #except:
                        #    printAndLog('Line ' + str(row) + '. Current stock name "' + str(c) + '" seems to be wrong. Skipped ...')
                        #    pass
            else:
                printAndLog('*** ERROR: Line ' + str(row) + '. Unable to determine stock date. Finding next "Ticker" row.')
try:
    wb.save(filename = r'Stocks.xlsx')
except:
    pass
printAndLog('DEBUG: DONE!')
logfile.close()